# In[]:

import os
import pandas as pd 
import openpyxl

#Certificatie
Path_C = r"P:\2022\22218 WNZ diverse vakken LN 2023\V1\07 Laboratorium\2 Certificaten\Excel" 
Path_Save = r"P:\2022\22218 WNZ diverse vakken LN 2023\V1\07 Laboratorium\2 Certificaten\3.xlsx"
#Extract the columns where the parameters concentrations 
# Are located
Monsters_MhPoly = []
Org_Stof =[]
PFOS = []
PFOA = []
EtFOSAA = []
MeFOSAA = []
HAP = []
Name_Par = []

for filename in os.listdir(Path_C):
    f = os.path.join(Path_C, filename)

    # Load the Excel file
    workbook = openpyxl.load_workbook(f)
    # Select the active worksheet
    worksheet = workbook.active


    #Restrict the area to only PFAS results

    WorkArea_Rows = []

    # Loop through all rows in the worksheet
    for row_index, row in enumerate(worksheet.iter_rows(), start=1):

        # Loop through all cells in the row
        for cell in row:

            # Check if the cell contains the word "Parameters"
            if cell.value == "perfluorbutaanzuur (PFBA)":

                # If it does, add the row position to the list
                WorkArea_Rows.append(row_index)

            if cell.value == "som PFOS":

                # If it does, add the row position to the list
                WorkArea_Rows.append(row_index)
            
    Columns = []
    Columns_NormMons =[]
    # Iterate over the rows in the specified range
    for row in worksheet.iter_rows():
        for cell in row:
            # Check if the cell value contains "µg/kg ds"
            if cell.value == "µg/kg ds":
                # Get the column position of the cell
                Columns.append(cell.column-1)

    for row in worksheet.iter_rows(max_row=WorkArea_Rows[0]):
        for cell in row:
            # Check if the cell value contains "mg/kg ds"
            if cell.value == "mg/kg ds":
                # Get the column position of the cell
                Columns_NormMons.append(cell.column-1)

    # Convert the list to a set to remove duplicates
    Columns_NormMons = set(Columns_NormMons)
    # Convert the set back to a list
    Columns_NormMons = list(Columns_NormMons)
    Columns_NormMons.sort()

    # Convert the list to a set to remove duplicates
    Columns = set(Columns)
    # Convert the set back to a list
    Columns = list(Columns)
    Columns.sort()

    # Extract the name of the monsters used by MHPoly
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == "Projectomschrijving":
                for x in Columns:
                    next_cell = worksheet.cell(row=cell.row, column=x)
                    Monsters_MhPoly.append(next_cell.value)
            if cell.value == "Org.Stof.cor":
                for x in Columns_NormMons:
                    next_cell = worksheet.cell(row=cell.row, column=x)
                    Org_Stof.append(next_cell.value)
    # print(Monsters_MhPoly)
    #Now let's extract the values 
    # Create the pandas
    df = pd.DataFrame({})

    for row in worksheet.iter_rows(min_row=WorkArea_Rows[0], max_row=WorkArea_Rows[1]):
        for cell in row:
            #Check for the common parameters
            if cell.value == "SOM PFOS":
                for x in Columns:
                    next_cell = worksheet.cell(row=cell.row, column=x)
                    PFOS.append(next_cell.value)   
            if cell.value == "SOM PFOA":
                for x in Columns:
                    next_cell = worksheet.cell(row=cell.row, column=x)
                    PFOA.append(next_cell.value)   
            if cell.value == "EtFOSAA":
                for x in Columns:
                    next_cell = worksheet.cell(row=cell.row, column=x)
                    EtFOSAA.append(next_cell.value)   
            if cell.value == "MeFOSAA":
                for x in Columns:
                    next_cell = worksheet.cell(row=cell.row, column=x)
                    MeFOSAA.append(next_cell.value)   

    #Hoogste Andere PFAS 
    Com_Par = ["SOM PFOS","SOM PFOA","EtFOSAA","MeFOSAA","C08: PFOA","C08: PFOAv","C08: PFOS" 
            ,"C08: PFOSv"] 

    RowsOtherPFAS = []

    for row in worksheet.iter_rows(min_row=WorkArea_Rows[0], max_row=WorkArea_Rows[1]
                                , max_col=1):
        for cell in row:
            #Check for the common parameters
            if cell.value not in Com_Par:
                RowsOtherPFAS.append(cell.row)

    # Now let's extract the highest PFAS value for the remaining rows

    Useless_List = []
    Useless_List_2 = []

    for x in Columns:
        i = 0
        for row_num in RowsOtherPFAS:
            cell_value = worksheet.cell(row=row_num, column=x).value
            Useless_List.append(cell_value)
        # Filter out invalid entries
        filtered_list = [x for x in Useless_List if isinstance(x, float)]
        
        try:
        
            HAP.append(max(filtered_list))
            Useless_List = []
            filtered_list = []
            for row_num in RowsOtherPFAS:
                cell_value = worksheet.cell(row=row_num, column=x).value
                if cell_value == HAP[-1]:
                    i = i + 1
                    cell_value = worksheet.cell(row=row_num, column=1).value
                    Useless_List_2.append(cell_value)

            if len(Useless_List_2)>1:
                result = ', '.join(Useless_List_2)
                Name_Par.append(result)
            else:
                Name_Par.append(Useless_List_2[0]) 
            Useless_List_2 = []
            i = 0
                
        #Empty the lists 
        except:
                Name_Par.append("--")
                HAP.append("--")

        Useless_List = []
        filtered_list = []

    # print(f)


# In[]: 

df['Mengmonster'] = Monsters_MhPoly
df['Som PFOS (µg/kg ds)'] = PFOS
df['SOM PFOA (µg/kg ds)'] = PFOA
df['EtFOSAA (µg/kg ds)'] = EtFOSAA
df['MeFOSAA (µg/kg ds)'] = MeFOSAA
df['Hoogste andere PFAS'] = Name_Par
df ["Concentratie (µg/kg ds)"] = HAP



#In[]: 

df["Organische stof (%)"] = Org_Stof
# df.sort_values('Mengmonster',inplace=True)

#In[]

Corr = []
#We can re use, it is no longer needed! 
Columns = ['Som PFOS (µg/kg ds)',
 'SOM PFOA (µg/kg ds)',
 'EtFOSAA (µg/kg ds)',
 'MeFOSAA (µg/kg ds)',
 'Concentratie (µg/kg ds)']

#In[]: 
numeric_col = pd.to_numeric(df['Organische stof (%)'], errors='coerce')

#In[]: 

# apply mask only to numeric values
mask = pd.notnull(numeric_col) & (numeric_col > 10)

#In[]: 

if mask.any():
    
    df.loc[mask, Columns] = df.loc[mask, Columns].apply(pd.to_numeric, errors='coerce') \
                            .div(df.loc[mask, 'Organische stof (%)'], axis=0) \
                            .fillna(0)

    df.loc[mask, Columns] = df[Columns].apply(lambda x: x * 10)

else: 
    pass

#In[]

for index, row in df.iterrows():
    if row["Organische stof (%)"] and pd.notnull(row["Organische stof (%)"])> 10:
        Corr.append("Ja")
    else:
        Corr.append("Nee")

df["Gecorrigeerd voor org.stof"] = Corr

# In[]:
df.to_excel(Path_Save)
# In[]: 
