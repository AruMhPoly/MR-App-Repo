#In[]: 

import requests
import os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tkinter import messagebox
import zipfile



class OMEGA_API: 

    def __init__(self, ProjectNummer,Path_certificaten_save):
        self.ProjectNummer = ProjectNummer
        self.endpoint = "https://mylab.omegam.nl/ws/results/download.aspx"
        self.api_key = "afd4037c-f0d5-11ed-a05b-0242ac120003"
        self.Path_certificaten_save = Path_certificaten_save

    def DownloadCertificates(self):

        result_params = {
            "type": "results",
            "project": self.ProjectNummer
        }
        result_headers = {
            "X-API-KEY": self.api_key
        }

        result_response = requests.get(self.endpoint,
                                       params=result_params, 
                                       headers=result_headers)
        
        global f1 
        global f2
        f1 = os.path.join(self.Path_certificaten_save,"CERTIFICATEN","EXCEL")
        f2 = os.path.join(self.Path_certificaten_save,"CERTIFICATEN","PDF")

        if not os.path.exists(f1):
                os.makedirs(f1)

        if not os.path.exists(f2):
            os.makedirs(f2)

        if result_response.status_code == 200:
            with open(os.path.join(f1,self.ProjectNummer + ".xml"), "wb") as file:
                file.write(result_response.content)
        
        else:
            message = self.ProjectNummer + "bestaat niet in het database. Probeer nog eens!"
            messagebox.showwarning("Warning", message)


        # Download ZIP file with all certificates
        cert_params = {
            "type": "certificates",
            "project": self.ProjectNummer
        }

        cert_headers = {
            "X-API-KEY": self.api_key
        }

        cert_response = requests.get(self.endpoint, params=cert_params, headers=cert_headers)

        if cert_response.status_code == 200:
            with open(os.path.join(f2,self.ProjectNummer + ".zip"), "wb") as file:
                file.write(cert_response.content)
        else:
            message = self.ProjectNummer + "bestaat niet in het database. Probeer nog eens!"
            messagebox.showwarning("Warning", message)

        p1 = os.path.join(f1,self.ProjectNummer + ".xml")
        return p1

    def convert_xml_to_excel(self):
        # Parse the XML file
        tree = ET.parse(os.path.join(f1,self.ProjectNummer + ".xml"))
        root = tree.getroot()

        # Iterate over each worksheet
        for worksheet in root.iter('{urn:schemas-microsoft-com:office:spreadsheet}Worksheet'):
            worksheet_name = worksheet.attrib['{urn:schemas-microsoft-com:office:spreadsheet}Name']

            # Create a new Excel workbook for the current worksheet
            workbook = Workbook()
            sheet = workbook.active

            # Write XML data to Excel for the current worksheet
            for row in worksheet.iter('{urn:schemas-microsoft-com:office:spreadsheet}Row'):
                row_num = int(row.attrib['{urn:schemas-microsoft-com:office:spreadsheet}Index'])
                for cell in row.iter('{urn:schemas-microsoft-com:office:spreadsheet}Cell'):
                    cell_num = int(cell.attrib['{urn:schemas-microsoft-com:office:spreadsheet}Index'])
                    column_letter = get_column_letter(cell_num)
                    cell_value = cell.find('{urn:schemas-microsoft-com:office:spreadsheet}Data').text
                    sheet[column_letter + str(row_num)].value = cell_value

            # Save the Excel file for the current worksheet
            excel_file = f'{worksheet_name}.xlsx'
            f = os.path.join(f1,excel_file)
            workbook.save(f)

    def Unzip(self): 

        # Create a ZipFile object
        zip_ref = zipfile.ZipFile(os.path.join(f2,self.ProjectNummer + ".zip"), 'r')

        # Extract all files to the specified path
        zip_ref.extractall(f2)

        # Close the ZipFile object
        zip_ref.close()

    def delete(self): 
        for filename in os.listdir(f1):
            # Check if the file has a .xml extension
            if filename.endswith(".xml"):
                file_path = os.path.join(f1, filename)
                # Delete the file
                os.remove(file_path)

        for filename in os.listdir(f2):
            # Check if the file has a .xml extension
            if filename.endswith(".zip"):
                file_path = os.path.join(f2, filename)
                # Delete the file
                os.remove(file_path)

#In[]: 

T1 = OMEGA_API(ProjectNummer="22196V1-Theodorushaven Bergen op Zoom",Path_certificaten_save=r"C:\Python\MR_APP\TESTEN_BOZ")
xml = T1.DownloadCertificates()
T1.convert_xml_to_excel()
T1.Unzip()
T1.delete()
#In[]:
