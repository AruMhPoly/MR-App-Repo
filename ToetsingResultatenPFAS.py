# In[]:

import os
import pandas as pd 
import openpyxl

class PFAS: 

    def __init__(self,PathSave, Path_Certificaten,ProjectNummer):
        self.Path_Certificaten = Path_Certificaten
        self.PathSave = PathSave
        self.ProjectNummer = ProjectNummer

    def ResultatenPFAS(self):

        Monsters = []
        Org_Stof =[]
        CommonPFAS = ["som PFOS (0.7 factor)",
                    "som PFOA (0.7 factor)",
                    "EtPFOSAA (n-ethyl perfluoroctaansulfonamide acetaat)",
                    "MePFOSAA (n-methyl perfluoroctaansulfonamide acetaat)"]
        IgnorePar = ["PFOA lineair (perfluoroctaanzuur)",
                    "PFOA vertakt (perfluoroctaanzuur)",
                    "PFOS lineair (perfluoroctaansulfonzuur)",
                    "PFOS vertakt (perfluoroctaansulfonzuur)",]


        PFOS = []
        PFOA = []
        EtFOSAA = []
        MeFOSAA = []
        HAP = []
        Name_Par = []
        Monsters =[]
        Org_Stof = []
        ROWS_PFAS = []

        for filename in os.listdir(self.Path_Certificaten):
            if filename.endswith(".xlsx"):  
                f = os.path.join(self.Path_Certificaten, filename)
                # Load the Excel file
                workbook = openpyxl.load_workbook(f)
                # Select the active worksheet
                worksheet = workbook.active
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.value == 'Monsteromschrijving':
                            for column in range(cell.column + 2 , worksheet.max_column):
                                    Monsters.append(worksheet.cell(row=cell.row, column=column).value)
                        
                        elif cell.value == 'organische stof (gloeiverlies)':
                            for column in range(cell.column + 2 , worksheet.max_column):
                                    if worksheet.cell(row=cell.row, column=column).value is not None:
                                        try:
                                            Org_Stof.append(float(worksheet.cell(row=cell.row, column=column).value))
                                        except:
                                            Org_Stof.append(float(0))
                                    else: 
                                        Org_Stof.append(float(worksheet.cell(row=cell.row, column=column - 1).value))

                        elif cell.value == 'PFBA (perfluorbutaanzuur)':      
                            RowPFBA = cell.row

                        #First the most common PFAS values
                        elif cell.value == CommonPFAS[0]:
                            for column in range(cell.column + 2 , worksheet.max_column):
                                try:
                                    PFOS.append(float(worksheet.cell(row=cell.row, column=column).value))
                                except:
                                    PFOS.append(worksheet.cell(row=cell.row, column=column).value)
                        
                        elif cell.value == CommonPFAS[1]:
                            for column in range(cell.column + 2 , worksheet.max_column):
                                try:
                                    PFOA.append(float(worksheet.cell(row=cell.row, column=column).value))
                                except:
                                    PFOA.append(worksheet.cell(row=cell.row, column=column).value)

                        elif cell.value == CommonPFAS[2]:
                            for column in range(cell.column + 2 , worksheet.max_column):
                                try:
                                    EtFOSAA.append(float(worksheet.cell(row=cell.row, column=column).value))
                                except:
                                    EtFOSAA.append(worksheet.cell(row=cell.row, column=column).value)

                        elif cell.value == CommonPFAS[3]:
                            for column in range(cell.column + 2 , worksheet.max_column):
                                try:
                                    MeFOSAA.append(float(worksheet.cell(row=cell.row, column=column).value))
                                except:
                                    MeFOSAA.append(worksheet.cell(row=cell.row, column=column).value)
                                    
                #Let's find the highest other PFAS values
                for row in worksheet.iter_rows(min_row=RowPFBA,max_col=1):
                    for cell in row:
                        if cell.value not in CommonPFAS and cell.value not in IgnorePar:
                            ROWS_PFAS.append(cell.row)  

                
                for column in range(3 , worksheet.max_column):
                    #List to add all PFAS values
                    UL = [] 
                    # List to add the name of the parameters
                    UL2 = []
                    # Columns with PFAS values. 
                    UL3 = [] 
                    for row in worksheet.iter_rows(min_row=RowPFBA):                
                            if row[0].row in ROWS_PFAS:
                                for cell in row:
                                            if cell.column == column:           
                                                try:
                                                    UL.append(float(worksheet.cell(row=cell.row, column=column).value))
                                                    UL2.append(worksheet.cell(row=cell.row, column=1).value)
                                                except:
                                                    UL.append(worksheet.cell(row=cell.row, column=cell.column).value)
                                                    UL2.append(worksheet.cell(row=cell.row, column=1).value)
                                    
                    #Get Maximum value 
                    max_value = float('-inf')
                    max_indices = []
                    for i, value in enumerate(UL):
                                    if not isinstance(value, str) and value is not None:  # Ignore strings and None Values 
                                        if value > max_value:
                                            max_value = value
                                            max_indices = [i]
                                        elif value == max_value:
                                            max_indices.append(i)
                                        
                    # Concatenate the parameters that belong to the highest PFAS
                    n  = str()
                    UN = 0
                    z = 0
                    for x in max_indices:
                        UN = UN + 1
                        if UN == 1:
                            n = str(UL2[x])
                            z = UL[max_indices[0]]
                        else:
                            n = n + "-" + str(UL2[x])
                
                    Name_Par.append(n)
                    HAP.append(z)
            
                workbook.close()

        df = pd.DataFrame(columns=['Mengmonster','Som PFOS (µg/kg ds)',
                                    'SOM PFOA (µg/kg ds)','EtFOSAA (µg/kg ds)',
                                    'MeFOSAA (µg/kg ds)','Hoogste andere PFAS'
                                    ,"Concentratie (µg/kg ds)","Organische stof (%)","Gecorrigeerd?"])

        df['Mengmonster'] = Monsters
        df['Som PFOS (µg/kg ds)'] = PFOS
        df['SOM PFOA (µg/kg ds)'] = PFOA
        df['EtFOSAA (µg/kg ds)'] = EtFOSAA
        df['MeFOSAA (µg/kg ds)'] = MeFOSAA
        df['Hoogste andere PFAS'] = Name_Par
        df ["Concentratie (µg/kg ds)"] = HAP
        df["Organische stof (%)"] = Org_Stof

        #The sum can not be 0.1, then it means it is not repported above the threshold value
        df.replace({'Som PFOS (µg/kg ds)': 0.1, 'SOM PFOA (µg/kg ds)': 0.1}, 0, inplace=True)
        # Find rows where column D is greater than 10
        condition = df['Organische stof (%)'] > 10
        # Update values in columns A and B
        try:
            df.loc[condition, 'Som PFOS (µg/kg ds)',] = (df['Som PFOS (µg/kg ds)'] / df['Organische stof (%)']) * 10
            df.loc[condition, 'SOM PFOA (µg/kg ds)',] = (df['SOM PFOA (µg/kg ds)'] / df['Organische stof (%)']) * 10
            df.loc[condition, 'EtFOSAA (µg/kg ds)',] = (df['EtFOSAA (µg/kg ds)'] / df['Organische stof (%)']) * 10
            df.loc[condition, 'MeFOSAA (µg/kg ds)',] = (df['MeFOSAA (µg/kg ds)'] / df['Organische stof (%)']) * 10
            df.loc[condition, 'Concentratie (µg/kg ds)',] = (df['Concentratie (µg/kg ds)'] / df['Organische stof (%)']) * 10
        except:
            pass

        df["Gecorrigeerd?"] = df['Organische stof (%)'].apply(lambda x: 'Ja' if x > 10 else 'Nee')
        df.set_index('Mengmonster',inplace=True)
        df.replace(0,"--",inplace=True)
        Path_Save = os.path.join(self.PathSave, self.ProjectNummer + '_Output_PFAS.xlsx')
        df.to_excel(Path_Save)
        return Path_Save

#In[]: 

# PS =r"C:\Python\MR_APP\SGS\19135V1\TOETSINGEN"
# PC = r"C:\Python\MR_APP\SGS\19135V1\CERTIFICAAT"
# Test = PFAS(PathSave=PS,Path_Certificaten=PC,ProjectNummer="23121V1")
# Test.ResultatenPFAS()
#In[]