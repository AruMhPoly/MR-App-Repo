#In[]: 

import os
import pandas as pd 
import openpyxl
import datetime

#In[]: 

class SlufterToets:
    
    def __init__(self,Projectnummer,Path_Certificaten,Path_Botova,Path_PFAS,Path_Toetsingen):
        self.ProjectNummer = Projectnummer
        self.Path_Certificaten = Path_Certificaten
        self.Path_Botova = Path_Botova
        self.Path_PFAS = Path_PFAS
        self.Path_Toetsingen = Path_Toetsingen
        
        self.Template = r"C:\Python\MR_APP\MR-App-Repo\SlufterToets\SluftertoetsTemplate.xlsx"

    def RunTest(self):

        Beoordelingen = {"Altijd toepasbaar":"AW",
                "Klasse A":"A",
                "Klasse B":"B",
                "Nooit toepasbaar":"NT"}
        
        #Monsters and Beoordeling

        df1 = pd.read_excel(self.Path_Botova)
        df1 = df1.iloc[2:,:]
        df1 = df1.sort_values(by='Unnamed: 0')
        MonstersBoToVa = df1["Unnamed: 0"].tolist()
        Beoordeling = df1["Unnamed: 2"].tolist()

        # Empty lists


        As = [] 
        Cd = []
        Cr = []
        Cu = []
        Hg = []
        Pb = []
        Ni = []
        Zn = []
        naftaleen = []
        anthraceen = []
        fenanthreen = []
        fluorantheen = []
        benzaanthr = []
        chryseen = []
        benzoapyre = []
        benzoghipe = []
        benzokfluo = []
        indeno123p = []
        PCB28 = []
        PCB52 = []
        PCB101 = []
        PCB118 = []
        PCB138 = []
        PCB153 = []
        PCB180 = []
        OLIEFLG = []
        HEXACHLB =[]
        SOMDDT =  []
        SOMDDD =  []
        SOMDDE =  []
        Aldrin = []
        Dieldrin = []
        Endrin = []
        Telodrin = []
        Isodrin = []
        AHCH = []
        BHCH = []
        YHCH = []
        heptachloor = []
        heptachloorepoxide = []
        hexachloorbutadieen = []

        DS = []
        OS = []
        Lutum = []

        Monsters_PFAS = []
        Monsters = []

        Parameters = ["Monster","Toeetsordeel",
                      "Droge Stof", "Organische Stof",
                      "Lutum",
                    "As (S)","Cd (S)","Cr (S)","Cu (S)",
                    "Pb (S)","Hg (S)","Ni (S)","Zn (S)",
                    "naftaleen","anthraceen","fenanthreen",
                    "fluorantheen","benz(a)anthr",
                    "chryseen","benzo(a)pyre",
                    "benzo(ghi)pe","benzo(k)fluo",
                    "indeno(123)p",
                    "PCB28","PCB52","PCB101","PCB118","PCB138",
                    "PCB153","PCB180","OLIE+FL(G)",
                    "HEXACHLB",
                    "SOM DDT", "SOM DDD", "SOM DDE",
                    "Aldrin","Dieldrin",
                        "Endrin","Telodrin","Isodrin",
                    "A-HCH","B-HCH","Y-HCH",
                    "Heptachloor","heptachloorepoxide",
                    "hexachloorbutadieen"]

        #WorkDataFrame
        df = pd.DataFrame({})

        # iterate over files in the directory
        for filename in os.listdir(self.Path_Certificaten):          
            f = os.path.join(self.Path_Certificaten, filename)
            ROWS_PFAS = []
            ROWS_OS = []
            # Load the Excel file
            workbook = openpyxl.load_workbook(f)
            # Select the active worksheet
            worksheet = workbook.active
            # Iterate through the rows starting from row 7
            for row in worksheet.iter_rows(min_row=7):
                # Check if "µg/kg ds" is present in any cell of the row
                for cell in row:
                    if cell.value == "µg/kg ds":
                        if row[1].value not in Monsters_PFAS:
                            Monsters_PFAS.append(row[1].value)
                            ROWS_PFAS.append(cell.row)
                    elif cell.value == "mg/kg ds":
                        if row[1].value not in Monsters:
                            Monsters.append(row[1].value) 
                            ROWS_OS.append(cell.row)

            for column in worksheet.iter_cols():
                        for cell in column:
                            if cell.value == "INDAMPR-SL" and cell.row == 1:
                                for y in ROWS_OS:
                                    DS.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Org.Stof.cor" and cell.row == 1:
                                for y in ROWS_OS:
                                    OS.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "LUT-SMICRO" and cell.row == 1:
                                for y in ROWS_OS:
                                    Lutum.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "As (S)" and cell.row == 1:
                                for y in ROWS_OS:
                                    As.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Cd (S)" and cell.row == 1:
                                for y in ROWS_OS:
                                    Cd.append(worksheet.cell(row=y, column=cell.column).value)
                            
                            elif cell.value == "Cr (S)" and cell.row == 1:
                                for y in ROWS_OS:
                                    Cr.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Cu (S)" and cell.row == 1:
                                for y in ROWS_OS:
                                    Cu.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Hg (S)" and cell.row == 1:
                                for y in ROWS_OS:
                                    Hg.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Pb (S)" and cell.row == 1:
                                for y in ROWS_OS:
                                    Pb.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Ni (S)" and cell.row == 1:
                                for y in ROWS_OS:
                                    Ni.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Zn (S)" and cell.row == 1:
                                for y in ROWS_OS:
                                    Zn.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "naftaleen" and cell.row == 1:                                
                                for y in ROWS_OS:
                                    naftaleen.append(worksheet.cell(row=y, column=cell.column).value)
                            
                            elif cell.value == "anthraceen" and cell.row == 1:
                                for y in ROWS_OS:
                                    anthraceen.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "fenanthreen" and cell.row == 1:
                                for y in ROWS_OS:
                                    fenanthreen.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "fluorantheen" and cell.row == 1:
                                for y in ROWS_OS:
                                    fluorantheen.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "benz(a)anthr" and cell.row == 1:
                                for y in ROWS_OS:
                                    benzaanthr.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "chryseen" and cell.row == 1:
                                for y in ROWS_OS:
                                    chryseen.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "benzo(a)pyre" and cell.row == 1:
                                for y in ROWS_OS:
                                    benzoapyre.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "benzo(ghi)pe" and cell.row == 1:
                                for y in ROWS_OS:
                                    benzoghipe.append(worksheet.cell(row=y, column=cell.column).value)
                            
                            elif cell.value == "benzo(k)fluo" and cell.row == 1:
                                for y in ROWS_OS:
                                    benzokfluo.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "indeno(123)p" and cell.row == 1:
                                for y in ROWS_OS:
                                    indeno123p.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "PCB28" and cell.row == 1:
                                for y in ROWS_OS:
                                    PCB28.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "PCB52" and cell.row == 1:
                                for y in ROWS_OS:
                                    PCB52.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "PCB101" and cell.row == 1:
                                for y in ROWS_OS:
                                    PCB101.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "PCB118" and cell.row == 1:
                                for y in ROWS_OS:
                                    PCB118.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "PCB138" and cell.row == 1:
                                for y in ROWS_OS:
                                    PCB138.append(worksheet.cell(row=y, column=cell.column).value)
                            
                            elif cell.value == "PCB153" and cell.row == 1:
                                for y in ROWS_OS:
                                    PCB153.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "PCB180" and cell.row == 1:
                                for y in ROWS_OS:
                                    PCB180.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "OLIE+FL(G)" and cell.row == 1:
                                for y in ROWS_OS:
                                    OLIEFLG.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "HEXACHLB" and cell.row == 1:
                                for y in ROWS_OS:
                                    HEXACHLB.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "SOM DDT" and cell.row == 1:
                                for y in ROWS_OS:
                                    SOMDDT.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "SOM DDD" and cell.row == 1:
                                for y in ROWS_OS:
                                    SOMDDD.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "SOM DDE" and cell.row == 1:
                                for y in ROWS_OS:
                                    SOMDDE.append(worksheet.cell(row=y, column=cell.column).value)
                            
                            elif cell.value == "Aldrin" and cell.row == 1:
                                for y in ROWS_OS:
                                    Aldrin.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Dieldrin" and cell.row == 1:
                                for y in ROWS_OS:
                                    Dieldrin.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Endrin" and cell.row == 1:
                                for y in ROWS_OS:
                                    Endrin.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Telodrin" and cell.row == 1:
                                for y in ROWS_OS:
                                    Telodrin.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Isodrin" and cell.row == 1:
                                for y in ROWS_OS:
                                    Isodrin.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "A-HCH" and cell.row == 1:
                                for y in ROWS_OS:
                                    AHCH.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "B-HCH" and cell.row == 1:
                                for y in ROWS_OS:
                                    BHCH.append(worksheet.cell(row=y, column=cell.column).value)
                            
                            elif cell.value == "Y-HCH" and cell.row == 1:
                                for y in ROWS_OS:
                                    YHCH.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "Heptachloor" and cell.row == 1:
                                for y in ROWS_OS:
                                    heptachloor.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "C-HEPCHEPO" and cell.row == 1:
                                for y in ROWS_OS:
                                    heptachloorepoxide.append(worksheet.cell(row=y, column=cell.column).value)

                            elif cell.value == "HCHLBUTADI" and cell.row == 1:
                                for y in ROWS_OS:
                                    hexachloorbutadieen.append(worksheet.cell(row=y, column=cell.column).value)

                           
            workbook.close()

        #First the general parameters

        df[Parameters[0]] = Monsters
        df[Parameters[1]] = Beoordeling
        df[Parameters[2]] = DS
        df[Parameters[3]] = OS
        df[Parameters[4]] = Lutum
        df[Parameters[5]] = As
        df[Parameters[6]] = Cd
        df[Parameters[7]] = Cr
        df[Parameters[8]] = Cu
        df[Parameters[9]] = Pb
        df[Parameters[10]] = Hg
        df[Parameters[11]] = Ni
        df[Parameters[12]] = Zn
        df[Parameters[13]] = naftaleen
        df[Parameters[14]] = anthraceen
        df[Parameters[15]] = fenanthreen
        df[Parameters[16]] = fluorantheen
        df[Parameters[17]] = benzaanthr
        df[Parameters[18]] = chryseen
        df[Parameters[19]] = benzoapyre
        df[Parameters[20]] = benzoghipe
        df[Parameters[21]] = benzokfluo
        df[Parameters[22]] = indeno123p
        df[Parameters[23]] = PCB28
        df[Parameters[24]] = PCB52
        df[Parameters[25]] = PCB101
        df[Parameters[26]] = PCB118
        df[Parameters[27]] = PCB138
        df[Parameters[28]] = PCB153
        df[Parameters[29]] = PCB180
        df[Parameters[30]] = OLIEFLG
        df[Parameters[31]] = HEXACHLB
        df[Parameters[32]] = SOMDDT
        df[Parameters[33]] = SOMDDD
        df[Parameters[34]] = SOMDDE
        df[Parameters[35]] = Aldrin
        df[Parameters[36]] = Dieldrin
        df[Parameters[37]] = Endrin
        df[Parameters[38]] = Telodrin
        df[Parameters[39]] = Isodrin
        df[Parameters[40]] = AHCH
        df[Parameters[41]] = BHCH
        df[Parameters[42]] = YHCH
        df[Parameters[43]] = heptachloor
        df[Parameters[44]] = heptachloorepoxide
        df[Parameters[45]] = hexachloorbutadieen
        # Replace non-numeric entries with zeros
        df = df.apply(pd.to_numeric, errors='coerce').fillna(0)
        df[Parameters[0]] = Monsters
        df[Parameters[1]] = Beoordeling
        df = df.sort_values(by='Monster')

        # Replace non-numeric entries with zeros
        dfPFAS = pd.read_excel(self.Path_PFAS)
        dfPFAS = dfPFAS.sort_values(by='Mengmonster')
        dfPFAS = dfPFAS.apply(pd.to_numeric, errors='coerce').fillna(0)
        dfPFAS2 = dfPFAS[["EtFOSAA (µg/kg ds)","MeFOSAA (µg/kg ds)","Concentratie (µg/kg ds)"]]
        dfPFAS2 = dfPFAS2.apply(pd.to_numeric, errors='coerce').fillna(0)
        # Add a new column with the maximum value of each row
        dfPFAS2['MaxValue'] = dfPFAS.iloc[:,:].apply(lambda row: row.max(), axis=1)        
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(self.Template)
        sheet = workbook.active
        col = 2
        #Begin with replacing 

        for index, row in df.iterrows():
            col = col + 1
            sheet.cell(row=1, column=col, value=row["Monster"])
            sheet.cell(row=3, column=col, value=Beoordelingen[row["Toeetsordeel"]])
            
    
            if dfPFAS.loc[index,"SOM PFOA (µg/kg ds)"] < 0.8 and dfPFAS.loc[index,"Som PFOS (µg/kg ds)"] < 3.70:
                sheet.cell(row=5, column=col, value="JA")
            else: 
                sheet.cell(row=5, column=col, value="NEE")

            sheet.cell(row=9, column=col, value=row[df.columns[2]])
            sheet.cell(row=10, column=col, value=row[df.columns[3]])
            sheet.cell(row=11, column=col, value=row[df.columns[4]])
            sheet.cell(row=16, column=col, value=row[df.columns[5]])
            sheet.cell(row=17, column=col, value=row[df.columns[6]])
            sheet.cell(row=18, column=col, value=row[df.columns[7]])
            sheet.cell(row=19, column=col, value=row[df.columns[8]])
            sheet.cell(row=20, column=col, value=row[df.columns[10]])
            sheet.cell(row=21, column=col, value=row[df.columns[9]])
            sheet.cell(row=22, column=col, value=row[df.columns[11]])
            sheet.cell(row=23, column=col, value=row[df.columns[12]])
            sheet.cell(row=25, column=col, value=row[df.columns[13]])
            sheet.cell(row=26, column=col, value=row[df.columns[14]])
            sheet.cell(row=27, column=col, value=row[df.columns[15]])
            sheet.cell(row=28, column=col, value=row[df.columns[16]])
            sheet.cell(row=29, column=col, value=row[df.columns[17]])
            sheet.cell(row=30, column=col, value=row[df.columns[18]])
            sheet.cell(row=31, column=col, value=row[df.columns[19]])
            sheet.cell(row=32, column=col, value=row[df.columns[20]])
            sheet.cell(row=33, column=col, value=row[df.columns[21]])
            sheet.cell(row=34, column=col, value=row[df.columns[22]])
            sheet.cell(row=38, column=col, value=row[df.columns[23]])
            sheet.cell(row=39, column=col, value=row[df.columns[24]])
            sheet.cell(row=40, column=col, value=row[df.columns[25]])
            sheet.cell(row=41, column=col, value=row[df.columns[26]])
            sheet.cell(row=42, column=col, value=row[df.columns[27]])
            sheet.cell(row=43, column=col, value=row[df.columns[28]])
            sheet.cell(row=44, column=col, value=row[df.columns[29]])
            sheet.cell(row=47, column=col, value=row[df.columns[30]])
            sheet.cell(row=49, column=col, value=row[df.columns[31]])
            sheet.cell(row=52, column=col, value=row[df.columns[32]])
            sheet.cell(row=53, column=col, value=row[df.columns[33]])
            sheet.cell(row=54, column=col, value=row[df.columns[34]])
            sheet.cell(row=57, column=col, value=row[df.columns[35]])
            sheet.cell(row=58, column=col, value=row[df.columns[36]])
            sheet.cell(row=59, column=col, value=row[df.columns[37]])
            sheet.cell(row=61, column=col, value=row[df.columns[38]])
            sheet.cell(row=62, column=col, value=row[df.columns[39]])
            sheet.cell(row=64, column=col, value=row[df.columns[40]])
            sheet.cell(row=65, column=col, value=row[df.columns[41]])
            sheet.cell(row=66, column=col, value=row[df.columns[42]])
            sheet.cell(row=67, column=col, value=row[df.columns[43]])
            sheet.cell(row=68, column=col, value=row[df.columns[44]])
            sheet.cell(row=69, column=col, value=row[df.columns[45]])
            sheet.cell(row=72, column=col, value=dfPFAS.loc[index,"Som PFOS (µg/kg ds)"])
            sheet.cell(row=73, column=col, value=dfPFAS.loc[index,"SOM PFOA (µg/kg ds)"])
            sheet.cell(row=74, column=col, value=dfPFAS2.loc[index,"MaxValue"])
            
        workbook.save(os.path.join(self.Path_Toetsingen , self.ProjectNummer + "Sluftertoets.xlsx"))
        workbook.close()


#In[]: 


# Projectnummer = "WNZ_"
# Path_Certificate = r"P:\2023\23121 WNZ monitoring 2023\V1\09 Laboratorium\02 Certificaten\EXCEL"
# Path_BoToVa = r"P:\2023\23121 WNZ monitoring 2023\V1\09 Laboratorium\03 Toetsingen\EXCEL\WNZ_Output_BoToVa.xlsx"
# Path_PFAS = r"P:\2023\23121 WNZ monitoring 2023\V1\09 Laboratorium\03 Toetsingen\EXCEL\WNZ_Output_PFAS.xlsx"
# Path_Toetsingen = r"P:\2023\23121 WNZ monitoring 2023\V1\09 Laboratorium\03 Toetsingen\EXCEL"
# x = SlufterToets(Projectnummer=Projectnummer,
#                  Path_Certificaten=Path_Certificate,
#                  Path_Botova=Path_BoToVa,
#                  Path_PFAS=Path_PFAS,
#                  Path_Toetsingen=Path_Toetsingen)
# x.RunTest()


#In[]: 