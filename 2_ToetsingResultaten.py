# In[]:

import os
import pandas as pd 
import openpyxl

#In[]

#PATHS - ONLY INPUT REQUIRED

#General path for all Toetsingen
Path_Toetsingen= r'P:\2022\22218 WNZ diverse vakken LN 2023\V1\07 Laboratorium\3 Toetsingen\RA01\EXCEL'
#Path for only T3: To identify the parameters 
# That cause a bad soil quality
Path_T3 = r"P:\2022\22218 WNZ diverse vakken LN 2023\V1\07 Laboratorium\3 Toetsingen\RA01\EXCEL\Botova_1512122_T3.xlsx"
#Certificaten
Path_Certificaten = r'P:\2022\22218 WNZ diverse vakken LN 2023\V1\07 Laboratorium\2 Certificaten\RA01\EXCEL'
#In[]: 

#WorkDataFrame
df = pd.DataFrame({})
#Empty lists
Columns_Names = []
Monsters =[]
Classification = []
Toetsingen = []

# iterate over files in the directory
for filename in os.listdir(Path_Toetsingen):
    f = os.path.join(Path_Toetsingen, filename)
    # Load the Excel file
    workbook = openpyxl.load_workbook(f)
    # Select the active worksheet
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == "Monster":
                next_cell = worksheet.cell(row=cell.row + 1, column=cell.column)
                Monsters.append(next_cell.value)
            if cell.value == "Toetsoordeel":
                next_cell = worksheet.cell(row=cell.row, column=cell.column + 1)
                Classification.append(next_cell.value)
            if cell.value == "Toetsing":
                next_cell = worksheet.cell(row=cell.row, column=cell.column + 1)
                Toetsingen.append(next_cell.value)

    New_Column=Toetsingen[0].split("-")[0].replace(".","").replace(" ","")
    Columns_Names.append(New_Column)

    # Add a new column for the new "Toetsing"
    df["Monster"] = Monsters
    df[New_Column] = Classification
    #Empty lists
    Columns_Names = []
    Monsters =[]
    Classification = []
    Toetsingen = []

#### What parameters produce that the klasse is B or NoT. 

# Load the Toetsing T3
workbook = openpyxl.load_workbook(Path_T3)

# Select the first worksheet
worksheet = workbook.worksheets[0]

# Create a list to store the row positions containing "Parameters" that are in high concentrations

parameter_rows = []

# Loop through all rows in the worksheet
for row_index, row in enumerate(worksheet.iter_rows(), start=1):

    # Loop through all cells in the row
    for cell in row:

        # Check if the cell contains the word "Parameters"
        if cell.value == "Parameters":

            # If it does, add the row position to the list
            parameter_rows.append(row_index)

            # Check if the cell contains the word "Parameters"

        if cell.value == "Legenda":

            # If it does, add the row position to the list
            parameter_rows.append(row_index)
        
end_rows=[]

# #This will allow me to restrict my area to a monster at the time

for x in range(0,len(parameter_rows)):
    end_rows.append(parameter_rows[x]- 2)

#In[]:

results = []
Monsters= []
Exceeded_Parameters = []

# Iterate through the first rows 
for row in range(parameter_rows[0], parameter_rows[1]):
    # Iterate through each column in the row
    for column in range(1, worksheet.max_column + 1):
        # Get the value of the cell
        cell_value = worksheet.cell(row=row, column=column).value
        # Check if the value is "T.Oordel"
        if cell_value == "T.Oordeel":
            # Print the column number where "T.Oordel" appears
            FirstColumn = column 
            # Exit the loop, since we've found what we're looking for
            break

#In[]:

for x in range(len(parameter_rows)-1):

    # Iterate over the rows in the specified range
    for row in worksheet.iter_rows(min_row=parameter_rows[x], max_row=end_rows[x+1]):
        # Check if the row contains the value 'B' in the last column
        # print(row[-1].value)
        if row[FirstColumn-1].value == 'B':
            # If it does, add the value from the first column to the results list
            results.append(row[0].value)
        if row[FirstColumn-1].value == 'NoT':
            # If it does, add the value from the first column to the results list
            results.append(row[0].value)

    Exceeded_Parameters.append(','.join(results))
    results = []

# Add a new column for the results
df["Parameters Overschreden bij T3"] = Exceeded_Parameters

#In[]:

df.to_excel(os.path.join(Path_Toetsingen,'Output_Botova.xlsx'))

#In[]: 
