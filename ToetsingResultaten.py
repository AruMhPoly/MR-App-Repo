# In[]:

import os
import pandas as pd 
import openpyxl

#In[]

Path= r"C:\Python\MR_APP\MR-App-Repo\Toetsingen"

# In[]:
#WorkDataFrame
df = pd.DataFrame({})
#Empty lists
Columns_Names = []
Monsters =[]
Classification = []
Toetsingen = []

# In[]

# iterate over files in
# that directory
for filename in os.listdir(Path):
    f = os.path.join(Path, filename)
    # Load the Excel file
    workbook = openpyxl.load_workbook(f)
    # Select the active worksheet
    worksheet = workbook.active
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == "Monster":
                next_cell = worksheet.cell(row=cell.row, column=cell.column + 1)
                Monsters.append(next_cell.value)
            if cell.value == "Toetsoordeel":
                next_cell = worksheet.cell(row=cell.row, column=cell.column + 1)
                Classification.append(next_cell.value)
            if cell.value == "Toetsing":
                next_cell = worksheet.cell(row=cell.row, column=cell.column + 1)
                Toetsingen.append(next_cell.value)

    New_Column=Toetsingen[0].split("-")[0].replace(".","").replace(" ","")
    Columns_Names.append(New_Column)
    # Add a new column for the new "Toetsing"
    df["Monster"] = Monsters
    df[New_Column] = Classification
    #Empty lists
    Columns_Names = []
    Monsters =[]
    Classification = []
    Toetsingen = []
# In[]:

#### What parameters produce that the klasse is B or NoT. 

Path_T3 = "C:\Python\MR_APP\MR-App-Repo\Toetsingen\Botova_1449959 + 1449958 + 1449957 + 1449956_T3.xlsx"

# Load the workbook
workbook = openpyxl.load_workbook(Path_T3)

# Select the first worksheet
worksheet = workbook.worksheets[0]

# Create a list to store the row positions containing "Parameters"
parameter_rows = []

# Loop through all rows in the worksheet
for row_index, row in enumerate(worksheet.iter_rows(), start=1):

    # Loop through all cells in the row
    for cell in row:

        # Check if the cell contains the word "Parameters"
        if cell.value == "Parameters":

            # If it does, add the row position to the list
            parameter_rows.append(row_index)

#In[]

end_rows=[]

for x in range(0,len(parameter_rows)):
    end_rows.append(parameter_rows[x]- 2)

#In[]

#Load the workbook
workbook = openpyxl.load_workbook(Path_T3)

# Select the worksheet that you want to search
worksheet = workbook.active

# Create a list to store the results
results = []
Monster = []

# Iterate over the rows in the specified range
for row in worksheet.iter_rows(min_row=257, max_row=335):
    # Check if the row contains the value 'B' in the first column
    if row[-1].value == 'B':
        # If it does, add the value from the first column to the results list
        results.append(row[0].value)
    if row[-1].value == 'NoT':
        # If it does, add the value from the first column to the results list
        results.append(row[0].value)
    if row[-3].value == 'Monster':
        # If it does, add the value from the first column to the results list
        Monster.append(row[-2].value)

#In[]:

Exceeded_Parameters = []

Exceeded_Parameters.append(','.join(results))
#In[]

