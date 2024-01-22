# In[]:

import os
import pandas as pd 
import openpyxl

class Botova:
    def __init__(self,Path_Toetsingen,ProjectNummer):

        self.Path_Toetsingen = Path_Toetsingen
        self.ProjectNummer = ProjectNummer

    def ResultatenBotova(self):

        #WorkDataFrame
        df = pd.DataFrame({})
        #Empty lists
        Columns_Names = []
        Monsters =[]
        Classification = []
        Monsters_Temporal = []
        # iterate over files in the directory
        for filename in os.listdir(self.Path_Toetsingen):
            f = os.path.join(self.Path_Toetsingen, filename)
            # Load the Excel file
            workbook = openpyxl.load_workbook(f)
            # Select the active worksheet
            worksheet = workbook["Sheet1"]
            UL = []
            UL2 = []
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == 'Monsteromschrijving\n              ':
                        # Iterate through the columns where Monsteromschrijving appears     
                        for column in range(cell.column + 2 , worksheet.max_column + 1):
                            UL.append(worksheet.cell(row=cell.row, column=column).value)
                            UL2.append(worksheet.cell(row=cell.row + 2, column=column).value)
                        break

            '''
            Since PFAs does not have a monter conclusie I need to filter
            the results so I get None for PFAS. This is not actually necesary
            because None is expresed as an empty cell once it's saved as an Excel file. 
            
            '''
            for x in range(len(UL)):
                if UL[x] is not None: 
                    Monsters_Temporal.append(UL[x])
                    if UL2[x] is not None:
                        Classification.append(UL2[x])
                    else: 
                        Classification.append(str())
            
            x =filename.split("_")[0]
            for i in range(len(Monsters_Temporal)):
                Columns_Names.append(x)
            Monsters.extend(Monsters_Temporal)
            Monsters_Temporal = []

        df["Monster"] = Monsters
        df["Toetsing"] = Columns_Names
        df['Classification'] = Classification
        df2 = df.pivot_table(index='Monster', columns='Toetsing', values='Classification', aggfunc=lambda x: ', '.join(x))
        workbook.close()

        Temporal_Exceeded_Parameters = [] 
        Exceeded_Parameters = []
        Monsters_Temporal = []
        Concetratie = []
        results = []
        
        for filename in os.listdir(self. Path_Toetsingen):
            f = os.path.join(self.Path_Toetsingen, filename)
            x =filename.split("_")[0]
            if x == "T3":
                #### What parameters produce that the klasse is B or NoT. 
                workbook = openpyxl.load_workbook(f)
                # Select the first worksheet
                worksheet = workbook.worksheets[0]
                # Create a list to store the row positions containing "Parameters" that are in high concentrations
                parameter_rows = []
                # Loop through all rows in the worksheet
                for row_index, row in enumerate(worksheet.iter_rows(), start=1):
                    # Loop through all cells in the row
                    for cell in row:
                        # Check if the cell contains the word "Parameters"
                        if cell.value in ['B','NoT']:
                            # If it does, add the row position to the list                    results = []
                            Concentratie = []
                            results.append(row[0].value)                
                            Concentratie.append(worksheet.cell(row=cell.row , column=cell.column - 2).value + " mg/kg ds ") 
                            # 6 is where monsteromschrijving is located
                            Monsters_Temporal.append(worksheet.cell(row=6 , column=cell.column - 2).value)
                            TemporalList = list(zip(results, Concentratie)) 
                            Temporal_Exceeded_Parameters.append(' - '.join([f"{x}: {y}" for x, y in TemporalList]))
                            TemporalList = []
                            results = []
                            Concentratie = []
                            Exceeded_Parameters.extend(Temporal_Exceeded_Parameters)
                            Temporal_Exceeded_Parameters = []

        if len(Exceeded_Parameters) > 0:
            df3 = pd.DataFrame({"Monster":Monsters_Temporal,"Parameters Overschreden bij T3":Exceeded_Parameters})
            dfM = df3.groupby('Monster').agg({
                'Parameters Overschreden bij T3': '- '.join
                })
            dff = pd.merge(df2, dfM, on='Monster',how='outer')

        else: 
            df2['Parameters Overschreden bij T3'] = [str()]*len(df2.iloc[:,0])
            dff = df2

        dff.to_excel(os.path.join(self.Path_Toetsingen,self.ProjectNummer + '_Output_BoToVa.xlsx'))

#In[]: 
# Path=  r"C:\Python\MR_APP\SGS\19135V1\TOETSINGEN\EXCEL"
# df = Botova(Path_Toetsingen=Path,ProjectNummer="TEST_").ResultatenBotova()

#In[]:

